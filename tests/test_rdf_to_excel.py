import json
from pathlib import Path
from sqlite3 import Blob

import pytest
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rdflib import SDO, SKOS, Literal, URIRef

from vocexcel.convert_084 import excel_to_rdf, rdf_to_excel
from vocexcel.utils import STATUSES, load_workbook, ShaclValidationError

def test_basic_084():
    RDF_FILE = Path(__file__).parent / "data" / "084_rdf.ttl"
    XL_FILE = RDF_FILE.with_suffix(".xlsx")
    rdf_to_excel(RDF_FILE, output_file=XL_FILE)

    g = excel_to_rdf(load_workbook(XL_FILE), output_format="graph")
    # g.serialize(destination=RDF_FILE.with_suffix(".2.ttl"), format="longturtle")

    assert (
        URIRef("https://linked.data.gov.au/def/induration-style"),
        SDO.status,
        URIRef(STATUSES["Experimental"]),
    ) in g

    assert (
        URIRef("https://linked.data.gov.au/def/induration-style"),
        SKOS.historyNote,
        Literal("Created from GA's exiting vocab for VocExcel testing"),
    ) in g

    assert (
        URIRef("https://linked.data.gov.au/def/induration-style/duricrust"),
        SKOS.relatedMatch,
        URIRef("http://example.com/fake"),
    ) in g

    assert (
        URIRef("https://linked.data.gov.au/def/induration-style/bauxitic_nodules"),
        SKOS.notation,
        Literal(
            "NO10",
            datatype=URIRef("https://linked.data.gov.au/def/induration-style/gaId"),
        ),
    ) in g

    XL_FILE.unlink(missing_ok=True)


def test_basic_084GA():
    RDF_FILE = Path(__file__).parent / "data" / "084GA_rdf.ttl"
    XL_FILE = RDF_FILE.with_suffix(".xlsx")
    rdf_to_excel(RDF_FILE, output_file=XL_FILE)

    g = excel_to_rdf(load_workbook(XL_FILE), output_format="graph")
    # g.serialize(destination=RDF_FILE.with_suffix(".2.ttl"), format="longturtle")
    assert (
        URIRef("https://linked.data.gov.au/def/induration-style"),
        SDO.status,
        URIRef(STATUSES["Experimental"]),
    ) in g

    assert (
        URIRef("https://linked.data.gov.au/def/induration-style"),
        SKOS.historyNote,
        Literal("Created from GA's exiting vocab for VocExcel testing"),
    ) in g

    XL_FILE.unlink(missing_ok=True)


def test_error_responses(capsys):
    RDF_FILE = Path(__file__).parent / "data" / "084_rdf_invalid.ttl"
    XL_FILE = RDF_FILE.with_suffix(".xlsx")

    # :completely_cemented_duricrust - no definition
    # :cs - no creator
    with pytest.raises(ShaclValidationError):
        rdf_to_excel(RDF_FILE, output_file=XL_FILE)
    Path(XL_FILE).unlink(missing_ok=True)

    with pytest.raises(ValueError):
        rdf_to_excel(RDF_FILE, output_file=XL_FILE.with_suffix(".xlsz"))  # 'python'

    rdf_to_excel(RDF_FILE, output_file=XL_FILE.with_suffix(".xlsz"), error_format="cmd")
    assert "If specifying an output_file, it must end with .xlsx" in capsys.readouterr().out

    j = rdf_to_excel(RDF_FILE, output_file=XL_FILE.with_suffix(".xlsz"), error_format="json")
    assert "If specifying an output_file, it must end with .xlsx" in j
    p = json.loads(j)
    assert "If specifying an output_file, it must end with .xlsx" in p["message"]

    Path(XL_FILE).unlink(missing_ok=True)

    with pytest.raises(ShaclValidationError):
        rdf_to_excel(RDF_FILE)  # 'python'

    rdf_to_excel(RDF_FILE, error_format="cmd")
    assert "RDF Validation Error" in capsys.readouterr().out

    j = rdf_to_excel(RDF_FILE, error_format="json")
    assert "RDF Validation Error" in j
    p = json.loads(j)
    assert "RDF Validation Error" in p["error"]

    Path(XL_FILE).unlink(missing_ok=True)


def test_return_formats():
    RDF_FILE = Path(__file__).parent / "data" / "084_rdf.ttl"
    XL_FILE = RDF_FILE.with_suffix(".xlsx")

    with pytest.raises(ValueError):
        rdf_to_excel(RDF_FILE, output_file=XL_FILE, output_format="graph")


    b = rdf_to_excel(RDF_FILE, output_file=XL_FILE, output_format="blob")
    assert isinstance(b, Workbook)

    XL_FILE.unlink(missing_ok=True)
    rdf_to_excel(RDF_FILE, output_file=XL_FILE, output_format="file")
    assert XL_FILE.exists()

    XL_FILE.unlink(missing_ok=True)