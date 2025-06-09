from pathlib import Path
from typing import Literal as TypeLiteral
from typing import Optional

from dateutil.parser import parse as date_parser
from openpyxl.styles import Alignment, Font
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pyshacl import validate as shacl_validate
from rdflib import DCTERMS, BNode, Graph, Literal, Namespace, URIRef
from rdflib.namespace import OWL, PROV, RDF, RDFS, SDO, SKOS, XSD

from vocexcel.convert_070 import (
    extract_additions_concept_properties as extract_additions_concept_properties_070,
)
from vocexcel.convert_070 import extract_collections as extract_collections_070
from vocexcel.convert_070 import extract_prefixes as extract_prefixes_070
from vocexcel.utils import (
    RDF_FILE_ENDINGS,
    STATUSES,
    VOCDERMODS,
    ConversionError,
    add_top_concepts,
    bind_namespaces,
    load_workbook,
    make_agent,
    make_iri,
    split_and_tidy_to_iris,
    split_and_tidy_to_strings,
    xl_hyperlink,
)

DATAROLES = Namespace("https://linked.data.gov.au/def/data-roles/")


def extract_prefixes(sheet: Worksheet):
    return extract_prefixes_070(sheet)


def extract_concept_scheme(
    sheet: Worksheet, prefixes, template_version="0.8.0"
) -> tuple[Graph, str]:
    iri_s = sheet["B3"].value
    title = sheet["B4"].value
    description = sheet["B5"].value
    created = sheet["B6"].value
    modified = sheet["B7"].value
    creator = sheet["B8"].value
    publisher = sheet["B9"].value
    custodian = sheet["B10"].value
    version = str(sheet["B11"].value).strip("'")
    history_note = sheet["B12"].value
    citation = sheet["B13"].value
    derived_from = sheet["B14"].value
    voc_der_mod = sheet["B15"].value
    themes = split_and_tidy_to_strings(sheet["B16"].value)
    status = sheet["B17"].value
    if template_version == "0.8.0.GA":
        catalogue_pid = sheet["B18"].value

    if iri_s is None:
        raise ConversionError(
            "Your vocabulary has no IRI. Please add it to the Concept Scheme sheet"
        )
    else:
        iri = make_iri(iri_s, prefixes)

    if title is None:
        raise ConversionError(
            "Your vocabulary has no title. Please add it to the Concept Scheme sheet"
        )

    if description is None:
        raise ConversionError(
            "Your vocabulary has no description. Please add it to the Concept Scheme sheet"
        )

    if created is None:
        raise ConversionError(
            "Your vocabulary has no created date. Please add it to the Concept Scheme sheet"
        )

    if modified is None:
        raise ConversionError(
            "Your vocabulary has no modified date. Please add it to the Concept Scheme sheet"
        )

    if creator is None:
        raise ConversionError(
            "Your vocabulary has no creator. Please add it to the Concept Scheme sheet"
        )

    if publisher is None:
        raise ConversionError(
            "Your vocabulary has no publisher. Please add it to the Concept Scheme sheet"
        )

    if history_note is None:
        raise ConversionError(
            "Your vocabulary has no History Note statement. Please add it to the Concept Scheme sheet"
        )

    # citation

    if derived_from is not None:
        if voc_der_mod is None:
            raise ConversionError(
                "If you supply a 'Derived From' value - IRI of another vocab - "
                "you must also supply a 'Derivation Mode' value"
            )

        if voc_der_mod not in VOCDERMODS:
            raise ConversionError(
                f"You have supplied a vocab derivation mode for your vocab of {voc_der_mod} but it is not recognised. "
                f"If supplied, it must be one of {', '.join(VOCDERMODS.keys())}"
            )

        derived_from = make_iri(derived_from, prefixes)

    # keywords

    if status is not None and status not in STATUSES:
        raise ConversionError(
            f"You have supplied a status for your vocab of {status} but it is not recognised. "
            f"If supplied, it must be one of {', '.join(STATUSES.keys())}"
        )

    if template_version == "0.8.0.GA":
        if catalogue_pid is None or not str(catalogue_pid).startswith(
            "https://pid.geoscience.gov.au/"
        ):
            raise ConversionError(
                "All GA vocabularies must have an eCat ID starting https://pid.geoscience.gov.au/dataset/..., assigned in the Concept Scheme metadata"
            )

    g = Graph(bind_namespaces="rdflib")
    g.add((iri, RDF.type, SKOS.ConceptScheme))
    g.add((iri, SKOS.prefLabel, Literal(title, lang="en")))
    g.add((iri, SKOS.definition, Literal(description, lang="en")))
    g.add((iri, SDO.dateCreated, Literal(created.date(), datatype=XSD.date)))
    g.add((iri, SDO.dateModified, Literal(modified.date(), datatype=XSD.date)))
    g += make_agent(creator, SDO.creator, prefixes, iri)
    g += make_agent(publisher, SDO.publisher, prefixes, iri)

    if custodian is not None:
        for _custodian in split_and_tidy_to_strings(custodian):
            g += make_agent(_custodian, DATAROLES.custodian, prefixes, iri)
            g.bind("DATAROLES", DATAROLES)

    if version is not None:
        g.add((iri, SDO.version, Literal(str(version))))
        g.add((iri, OWL.versionIRI, URIRef(iri + "/" + str(version))))

    g.add((iri, SKOS.historyNote, Literal(history_note)))

    if citation is not None:
        if str(citation).startswith("http"):
            val = Literal(citation, datatype=XSD.anyURI)
        else:
            val = Literal(citation)

        g.add((iri, SDO.citation, val))

    if derived_from is not None:
        qd = BNode()
        g.add((iri, PROV.qualifiedDerivation, qd))
        g.add((qd, PROV.entity, URIRef(derived_from)))
        g.add((qd, PROV.hadRole, URIRef(VOCDERMODS[voc_der_mod])))

    if themes is not None:
        for theme in themes:
            try:
                theme = make_iri(theme, prefixes)
            except ConversionError:
                theme = Literal(theme)
            g.add((iri, SDO.keywords, theme))

    if status is not None:
        g.add((iri, SDO.status, URIRef(STATUSES[status])))

    if template_version == "0.8.0.GA":
        g.add((iri, SDO.identifier, Literal(catalogue_pid, datatype=XSD.anyURI)))

    bind_namespaces(g, prefixes)
    g.bind("", Namespace(str(iri) + "/"))
    return g, iri


def extract_concepts(sheet: Worksheet, prefixes, cs_iri):
    g = Graph(bind_namespaces="rdflib")
    i = 4
    while True:
        # get values
        iri_s = sheet[f"A{i}"].value
        pref_label = sheet[f"B{i}"].value
        definition = sheet[f"C{i}"].value
        alt_labels = sheet[f"D{i}"].value
        narrower = sheet[f"E{i}"].value
        history_note = sheet[f"F{i}"].value
        citation = sheet[f"G{i}"].value
        is_defined_by = sheet[f"H{i}"].value
        status = sheet[f"I{i}"].value
        example = sheet[f"J{i}"].value
        image_url = sheet[f"K{i}"].value
        image_embedded = sheet[f"L{i}"].value

        # check values
        if iri_s is None:
            break

        iri = make_iri(iri_s, prefixes)

        if pref_label is None:
            raise ConversionError(
                f"You must provide a Preferred Label for Concept {iri_s}"
            )

        if definition is None:
            raise ConversionError(f"You must provide a Definition for Concept {iri_s}")

        if status is not None and status not in STATUSES:
            raise ConversionError(
                f"You have supplied a status for your Concept of {status} but it is not recognised. "
                f"If supplied, it must be one of {', '.join(STATUSES.keys())}"
            )

        if image_url is not None:
            if not image_url.startswith("http"):
                raise ConversionError(
                    "If supplied, an Image URL must start with 'http'"
                )

        # TODO: test embedded mage is not too large
        # image_embedded
        if image_embedded not in [None, "#VALUE!"]:
            raise ConversionError(
                "The Image Embedded colum, you must only insert images or leave it blank. "
                f"You have an unexpected value in Cell L{i}"
            )

        # ignore example Concepts
        if iri_s in [
            "http://example.com/earth-science",
            "http://example.com/atmospheric-science",
            "http://example.com/geology",
        ]:
            continue

        # create Graph
        g.add((iri, RDF.type, SKOS.Concept))
        g.add((iri, SKOS.inScheme, cs_iri))

        if "@" in pref_label:
            val, lang = pref_label.strip().split("@")
            g.add((iri, SKOS.prefLabel, Literal(val, lang=lang)))
        else:
            g.add((iri, SKOS.prefLabel, Literal(pref_label.strip(), lang="en")))
        g.add((iri, SKOS.definition, Literal(definition.strip(), lang="en")))

        if alt_labels is not None:
            for al in split_and_tidy_to_strings(alt_labels):
                g.add((iri, SKOS.altLabel, Literal(al, lang="en")))

        if narrower is not None:
            for n in split_and_tidy_to_iris(narrower, prefixes):
                g.add((iri, SKOS.narrower, n))
                g.add((n, SKOS.broader, iri))

        if history_note is not None:
            g.add((iri, SKOS.historyNote, Literal(history_note.strip())))

        if citation is not None:
            for _citation in split_and_tidy_to_strings(citation):
                g.add(
                    (iri, SDO.citation, Literal(_citation.strip(), datatype=XSD.anyURI))
                )

        if is_defined_by is not None:
            g.add((iri, RDFS.isDefinedBy, URIRef(is_defined_by.strip())))
        else:
            g.add((iri, RDFS.isDefinedBy, cs_iri))

        if status is not None:
            g.add((iri, SDO.status, URIRef(STATUSES[status])))

        if example is not None:
            g.add((iri, SKOS.example, Literal(str(example).strip())))

        if image_url is not None:
            g.add(
                (iri, SDO.image, Literal(str(image_url).strip(), datatype=XSD.anyURI))
            )

        if image_embedded == "#VALUE!":
            g.add((iri, SDO.image, Literal(f"Image at L{i}")))

        i += 1

    bind_namespaces(g, prefixes)
    return g


def extract_collections(sheet: Worksheet, prefixes, cs_iri):
    return extract_collections_070(sheet, prefixes, cs_iri)


def extract_additions_concept_properties(sheet: Worksheet, prefixes):
    return extract_additions_concept_properties_070(sheet, prefixes)


def excel_to_rdf(
    wb: Workbook,
    output_file_path: Optional[Path] = None,
    output_format: TypeLiteral[
        "longturtle", "turtle", "xml", "json-ld", "graph"
    ] = "longturtle",
    template_version="0.8.0",
):
    if template_version not in ["0.8.0", "0.8.0.GA"]:
        raise ValueError(
            f"This converter can only handle templates with versions 0.8.0 or 0.8.0.GA, not {template_version}"
        )

    prefixes = extract_prefixes(wb["Prefixes"])
    cs, cs_iri = extract_concept_scheme(
        wb["Concept Scheme"], prefixes, template_version
    )
    cons = extract_concepts(wb["Concepts"], prefixes, cs_iri)
    cols = extract_collections(wb["Collections"], prefixes, cs_iri)
    extra = extract_additions_concept_properties(
        wb["Additional Concept Properties"], prefixes
    )

    g = cs + cons + cols + extra
    g = add_top_concepts(g)
    g.bind("cs", cs_iri)

    # validate the RDF file
    shacl_graph = Graph().parse(Path(__file__).parent / "vocpub-5.1.ttl")
    v = shacl_validate(g, shacl_graph=shacl_graph, allow_warnings=True)
    if not v[0]:
        raise ConversionError(v[2])

    if output_file_path is not None:
        g.serialize(destination=str(output_file_path), format="longturtle")
    else:  # print to std out
        if output_format == "graph":
            return g
        else:
            return g.serialize(format=output_format)


def rdf_to_excel(
    rdf_file: Path, output_file_path: Optional[Path] = None, template_version="0.8.0"
):
    # value checkers
    if not rdf_file.name.endswith(tuple(RDF_FILE_ENDINGS.keys())):
        raise ValueError(
            "Files for conversion to Excel must end with one of the RDF file formats: '{}'".format(
                "', '".join(RDF_FILE_ENDINGS.keys())
            )
        )

    if output_file_path is not None:
        if not output_file_path.suffix == ".xlsx":
            raise ValueError(
                "If specifying an output_file_path, it must end with .xlsx"
            )

    if template_version not in ["0.8.0", "0.8.0.GA"]:
        raise ValueError(
            f"This converter can only handle templates with versions 0.8.0 or 0.8.0.GA, not {template_version}"
        )

    # load the RDF file
    g = Graph().parse(str(rdf_file), format=RDF_FILE_ENDINGS[rdf_file.suffix])
    g.bind("ex", "http://example.com/")
    ns = g.namespace_manager

    # validate the RDF file
    shacl_graph = Graph().parse(Path(__file__).parent / "vocpub-5.1.ttl")
    v = shacl_validate(g, shacl_graph=shacl_graph, allow_warnings=True)
    if not v[0]:
        raise ConversionError(v[2])

    # load the template
    fn = (
        "VocExcel-template-080-GA.xlsx"
        if template_version == "0.8.0.GA"
        else "VocExcel-template-080.xlsx"
    )
    wb = load_workbook(Path(__file__).parent.parent / "templates" / fn)

    # Concept Scheme
    ws = wb["Concept Scheme"]
    cs_iri = g.value(predicate=RDF.type, object=SKOS.ConceptScheme)
    ws["B3"] = cs_iri
    ws["B4"] = g.value(subject=cs_iri, predicate=SKOS.prefLabel)
    ws["B5"] = g.value(subject=cs_iri, predicate=SKOS.definition)
    ws["B5"].alignment = Alignment(wrap_text=True)
    ws["B6"] = date_parser(
        str(
            g.value(subject=cs_iri, predicate=SDO.dateCreated)
            or g.value(subject=cs_iri, predicate=DCTERMS.created)
        )
    )
    ws["B7"] = date_parser(
        str(
            g.value(subject=cs_iri, predicate=SDO.dateModified)
            or g.value(subject=cs_iri, predicate=DCTERMS.modified)
        )
    )
    xl_hyperlink(
        ws["B8"],
        g.value(subject=cs_iri, predicate=SDO.creator)
        or g.value(subject=cs_iri, predicate=DCTERMS.creator),
    )
    ws["B9"] = g.value(subject=cs_iri, predicate=SDO.publisher) or g.value(
        subject=cs_iri, predicate=DCTERMS.publisher
    )
    # custodian
    for o in g.objects(subject=cs_iri, predicate=PROV.qualifiedAttribution):
        for p, o2 in g.predicate_objects(subject=o):
            if p == DATAROLES.custodian:
                ws["B10"] = str(o2)
    ws["B11"] = g.value(subject=cs_iri, predicate=SDO.version) or g.value(
        subject=cs_iri, predicate=OWL.versionInfo
    )
    ws["B12"] = g.value(subject=cs_iri, predicate=SKOS.historyNote)
    ws["B13"] = g.value(subject=cs_iri, predicate=SDO.citation)
    for o in g.objects(subject=cs_iri, predicate=PROV.qualifiedDerivation):
        for p, o2 in g.predicate_objects(subject=o):
            if p == PROV.entity:
                ws["B14"] = str(o2)
            if p == PROV.hadRole:
                for k, v in VOCDERMODS.items():
                    if v == str(o2):
                        ws["B15"] = k
    ws["B16"] = ", ".join(
        [str(x) for x in g.objects(subject=cs_iri, predicate=SDO.keywords)]
    )
    ws["B17"] = str(g.value(subject=cs_iri, predicate=SDO.status)).split("/")[-1]
    if template_version == "0.8.0.GA":
        ws["B18"] = str(g.value(subject=cs_iri, predicate=SDO.identifier))

    # Concepts
    ws = wb["Concepts"]
    r = 4
    cs = sorted(list(g.subjects(predicate=RDF.type, object=SKOS.Concept)))
    for c in cs:
        xl_hyperlink(ws[f"A{r}"], ns.curie(c))
        ws[f"B{r}"] = g.value(subject=c, predicate=SKOS.prefLabel)
        ws[f"B{r}"].font = Font(size=14)
        ws[f"C{r}"] = g.value(subject=c, predicate=SKOS.definition)
        ws[f"C{r}"].alignment = Alignment(wrap_text=True)
        ws[f"C{r}"].font = Font(size=14)

        alt_labels = []
        for alt_label in g.objects(subject=c, predicate=SKOS.altLabel):
            alt_labels.append(alt_label)
        ws[f"D{r}"] = ",\n".join(alt_labels)
        ws[f"D{r}"].font = Font(size=14)

        for s, o in g.subject_objects(SKOS.broader):
            g.add((o, SKOS.narrower, s))
        narrowers = []
        for narrower in g.objects(subject=c, predicate=SKOS.narrower):
            narrowers.append(narrower)
        ws[f"E{r}"] = ",\n".join([ns.curie(x) for x in narrowers])
        ws[f"E{r}"].font = Font(size=14)

        hn = g.value(subject=c, predicate=SKOS.historyNote)
        if hn is not None:
            ws[f"F{r}"] = hn
            ws[f"F{r}"].font = Font(size=14)

        cit = g.value(subject=c, predicate=SDO.citation)
        if cit is not None:
            ws[f"G{r}"] = hn
            ws[f"G{r}"].font = Font(size=14)

        is_defined_by = g.value(subject=c, predicate=RDFS.isDefinedBy)
        if is_defined_by != "" and is_defined_by != cs_iri:
            xl_hyperlink(ws[f"H{r}"], is_defined_by)

        status = g.value(subject=c, predicate=SDO.status)
        if status is not None:
            for k, v in STATUSES.items():
                if v == str(status):
                    ws[f"I{r}"] = k
                    ws[f"I{r}"].font = Font(size=14)

        eg = g.value(subject=c, predicate=SKOS.example)
        if eg is not None:
            ws[f"J{r}"] = eg
            ws[f"J{r}"].font = Font(size=14)

        img = g.value(subject=c, predicate=SDO.image)
        if img is not None:
            xl_hyperlink(ws[f"K{r}"], img)

        r += 1

    r = 4
    cs = sorted(list(g.subjects(predicate=RDF.type, object=SKOS.Concept)))
    for c in cs:
        # Concepts - Additional Properties
        if (
            c,
            SKOS.relatedMatch
            | SKOS.closeMatch
            | SKOS.exactMatch
            | SKOS.narrowMatch
            | SKOS.broadMatch
            | SKOS.notation,
            None,
        ) in g:
            ws = wb["Additional Concept Properties"]
            xl_hyperlink(ws[f"A{r}"], ns.curie(c))

            fill_cell_with_list_of_curies(f"B{r}", ws, g, c, SKOS.relatedMatch)
            fill_cell_with_list_of_curies(f"C{r}", ws, g, c, SKOS.closeMatch)
            fill_cell_with_list_of_curies(f"D{r}", ws, g, c, SKOS.exactMatch)
            fill_cell_with_list_of_curies(f"E{r}", ws, g, c, SKOS.narrowMatch)
            fill_cell_with_list_of_curies(f"F{r}", ws, g, c, SKOS.broadMatch)

            notations = []
            datatypes = []
            for i in g.objects(subject=c, predicate=SKOS.notation):
                notations.append(str(i))
                datatypes.append(i.datatype)
            ws[f"G{r}"] = ",\n".join(notations)
            ws[f"G{r}"].font = Font(size=14)
            ws[f"H{r}"] = ",\n".join(datatypes)
            ws[f"H{r}"].font = Font(size=14)

            r += 1

    # Collections

    # Namespaces
    ws = wb["Prefixes"]
    common_prefixes = [
        "ex",
        "brick",
        "csvw",
        "dc",
        "dcat",
        "dcmitype",
        "dcterms",
        "dcam",
        "doap",
        "foaf",
        "geo",
        "odrl",
        "org",
        "prof",
        "prov",
        "qb",
        "schema",
        "sh",
        "skos",
        "sosa",
        "ssn",
        "time",
        "vann",
        "void",
        "wgs",
        "owl",
        "rdf",
        "rdfs",
        "xsd",
        "xml",
    ]
    r = 4
    for pre, ns in g.namespaces():
        if pre not in common_prefixes:
            ws[f"A{r}"] = pre
            ws[f"A{r}"].font = Font(size=14)
            xl_hyperlink(ws[f"B{r}"], ns)
            r += 1

    # save the output
    if output_file_path is None:
        wb.save(str(rdf_file.with_suffix(".xlsx")))
    else:
        wb.save(str(output_file_path))


def fill_cell_with_list_of_curies(
    cell_id: str, ws: Worksheet, g: Graph, subj: URIRef, pred: URIRef
):
    xs = []
    for x in g.objects(subject=subj, predicate=pred):
        xs.append(x)
    ws[cell_id] = ",\n".join([g.namespace_manager.curie(z) for z in xs])
    ws[cell_id].font = Font(size=14)
